# app.py - PostgreSQL Only Version

import os
import sys
import io
import csv
from flask import Flask, render_template, request, redirect, url_for, flash, session, Response, jsonify, send_from_directory, abort
from datetime import datetime, timedelta, date
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename
from functools import wraps
import secrets
import re
import psycopg
from psycopg.rows import dict_row
from psycopg.errors import UniqueViolation
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from decimal import Decimal, InvalidOperation

# Carregar variáveis de ambiente do arquivo .env se existir
try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # python-dotenv não está instalado, usar apenas variáveis de ambiente do sistema

# Inicializa a aplicação Flask
app = Flask(__name__)
# Usa SECRET_KEY do ambiente em produção; gera uma chave temporária caso não definida
app.secret_key = os.getenv('SECRET_KEY', secrets.token_hex(16))  # Chave secreta para sessões

# Configuração de uploads
UPLOAD_FOLDER = 'uploads'
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER


# --- Configuração e Inicialização do Banco de Dados ---
DATABASE_URL = os.getenv('DATABASE_URL')

if not DATABASE_URL:
    raise RuntimeError('DATABASE_URL não configurada. Defina a variável de ambiente.')

def get_db():
    """Abre uma nova conexão com o banco de dados PostgreSQL."""
    # Em ambientes gerenciados (ex.: Render), forçar SSL se não especificado
    db_url = DATABASE_URL
    if 'sslmode=' not in db_url and 'localhost' not in db_url and '127.0.0.1' not in db_url:
        separator = '&' if '?' in db_url else '?'
        db_url = f"{db_url}{separator}sslmode=require"
    
    conn = psycopg.connect(db_url)
    return conn

def init_db():
    """Inicializa o banco de dados PostgreSQL e cria as tabelas se não existirem."""
    with app.app_context():
        conn = get_db()
        cur = conn.cursor()

        # Tabela de usuários para autenticação
        cur.execute('''
            CREATE TABLE IF NOT EXISTS usuarios (
                id SERIAL PRIMARY KEY,
                nome TEXT NOT NULL,
                email TEXT UNIQUE NOT NULL,
                senha TEXT NOT NULL,
                tipo TEXT DEFAULT 'operador',
                nivel TEXT DEFAULT 'Operador',
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Tabela de clientes
        cur.execute('''
            CREATE TABLE IF NOT EXISTS clientes (
                id SERIAL PRIMARY KEY,
                nome TEXT NOT NULL,
                cpf_cnpj TEXT UNIQUE NOT NULL,
                rg TEXT NOT NULL,
                email TEXT,
                telefone TEXT,
                telefone_secundario TEXT,
                chave_pix TEXT NOT NULL,
                endereco TEXT NOT NULL,
                cidade TEXT NOT NULL,
                estado TEXT NOT NULL,
                cep TEXT NOT NULL,
                referencia TEXT NOT NULL,
                telefone_referencia TEXT NOT NULL,
                endereco_referencia TEXT NOT NULL,
                observacoes TEXT,
                empresa TEXT NOT NULL,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP
            )
        ''')

        # Tabela de documentos de clientes
        cur.execute('''
            CREATE TABLE IF NOT EXISTS documentos (
                id SERIAL PRIMARY KEY,
                cliente_id INTEGER NOT NULL,
                nome_ficheiro TEXT NOT NULL,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                CONSTRAINT fk_documentos_cliente FOREIGN KEY (cliente_id) REFERENCES clientes (id) ON DELETE CASCADE
            )
        ''')

        # Tabela de cobranças
        cur.execute('''
            CREATE TABLE IF NOT EXISTS cobrancas (
                id SERIAL PRIMARY KEY,
                cliente_id INTEGER NOT NULL,
                descricao TEXT,
                valor_original DOUBLE PRECISION NOT NULL,
                valor_pago DOUBLE PRECISION DEFAULT 0,
                multa DOUBLE PRECISION DEFAULT 0,
                juros DOUBLE PRECISION DEFAULT 0,
                desconto DOUBLE PRECISION DEFAULT 0,
                valor_total DOUBLE PRECISION,
                taxa_juros DOUBLE PRECISION DEFAULT 0,
                data_vencimento DATE NOT NULL,
                data_pagamento DATE,
                status TEXT DEFAULT 'Pendente',
                forma_pagamento TEXT,
                numero_parcelas INTEGER DEFAULT 1,
                parcela_atual INTEGER DEFAULT 1,
                tipo_cobranca TEXT DEFAULT 'Única',
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                CONSTRAINT fk_cobrancas_cliente FOREIGN KEY (cliente_id) REFERENCES clientes (id) ON DELETE CASCADE
            )
        ''')

        # Adicionar coluna taxa_juros se não existir
        cur.execute('''
            ALTER TABLE cobrancas 
            ADD COLUMN IF NOT EXISTS taxa_juros DOUBLE PRECISION DEFAULT 0
        ''')
        
        # Garantir que a coluna valor_pago existe
        cur.execute('''
            ALTER TABLE cobrancas 
            ADD COLUMN IF NOT EXISTS valor_pago DOUBLE PRECISION DEFAULT 0
        ''')
        
        # Adicionar coluna multa_manual para parcelas se não existir
        cur.execute('''
            ALTER TABLE parcelas 
            ADD COLUMN IF NOT EXISTS multa_manual DOUBLE PRECISION
        ''')
        
        # Remover coluna multa_manual da tabela cobrancas (não será mais usada)
        cur.execute('''
            ALTER TABLE cobrancas 
            DROP COLUMN IF EXISTS multa_manual
        ''')

        # Tabela de pagamentos individuais
        cur.execute('''
            CREATE TABLE IF NOT EXISTS pagamentos (
                id SERIAL PRIMARY KEY,
                cobranca_id INTEGER NOT NULL,
                valor_pago DOUBLE PRECISION NOT NULL,
                data_pagamento DATE NOT NULL DEFAULT CURRENT_DATE,
                observacao TEXT,
                forma_pagamento TEXT,
                usuario_id INTEGER,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                CONSTRAINT fk_pagamentos_cobranca FOREIGN KEY (cobranca_id) REFERENCES cobrancas (id) ON DELETE CASCADE,
                CONSTRAINT fk_pagamentos_usuario FOREIGN KEY (usuario_id) REFERENCES usuarios (id)
            )
        ''')

        # Tabela de histórico de pagamentos (mantida para compatibilidade)
        cur.execute('''
            CREATE TABLE IF NOT EXISTS historico_pagamentos (
                id SERIAL PRIMARY KEY,
                cobranca_id INTEGER NOT NULL,
                cliente_id INTEGER NOT NULL,
                valor_pago DOUBLE PRECISION NOT NULL,
                data_pagamento TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                forma_pagamento TEXT,
                observacoes TEXT,
                usuario_id INTEGER,
                CONSTRAINT fk_hist_cobranca FOREIGN KEY (cobranca_id) REFERENCES cobrancas (id) ON DELETE CASCADE,
                CONSTRAINT fk_hist_cliente FOREIGN KEY (cliente_id) REFERENCES clientes (id) ON DELETE CASCADE,
                CONSTRAINT fk_hist_usuario FOREIGN KEY (usuario_id) REFERENCES usuarios (id)
            )
        ''')

        # Tabela de notificações enviadas
        cur.execute('''
            CREATE TABLE IF NOT EXISTS notificacoes (
                id SERIAL PRIMARY KEY,
                cliente_id INTEGER NOT NULL,
                cobranca_id INTEGER NOT NULL,
                tipo TEXT NOT NULL,
                mensagem TEXT,
                status TEXT DEFAULT 'Enviada',
                data_envio TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                CONSTRAINT fk_notif_cliente FOREIGN KEY (cliente_id) REFERENCES clientes (id) ON DELETE CASCADE,
                CONSTRAINT fk_notif_cobranca FOREIGN KEY (cobranca_id) REFERENCES cobrancas (id) ON DELETE CASCADE
            )
        ''')


        # Tabela de parcelas
        cur.execute('''
            CREATE TABLE IF NOT EXISTS parcelas (
                id SERIAL PRIMARY KEY,
                cobranca_id INTEGER NOT NULL,
                numero_parcela INTEGER NOT NULL,
                valor DOUBLE PRECISION NOT NULL,
                data_vencimento DATE NOT NULL,
                status TEXT DEFAULT 'Pendente',
                valor_pago DOUBLE PRECISION DEFAULT 0,
                data_pagamento DATE,
                forma_pagamento TEXT,
                observacoes TEXT,
                criado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                atualizado_em TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                CONSTRAINT fk_parcelas_cobranca FOREIGN KEY (cobranca_id) REFERENCES cobrancas (id) ON DELETE CASCADE
            )
        ''')

        # Tabela de configurações do sistema
        cur.execute('''
            CREATE TABLE IF NOT EXISTS configuracoes (
                id SERIAL PRIMARY KEY,
                chave TEXT UNIQUE NOT NULL,
                valor TEXT,
                descricao TEXT
            )
        ''')

        # Inserir configurações padrão (ignora se já existir)
        default_configs = [
            ('taxa_juros_mensal', '2.0', 'Taxa de juros mensal (%)'),
            ('taxa_multa', '10.0', 'Taxa de multa por atraso (%)'),
            ('dias_tolerancia', '3', 'Dias de tolerância antes de aplicar multa'),
            ('envio_automatico', 'true', 'Ativar envio automático de notificações'),
            ('dias_aviso_vencimento', '3', 'Dias antes do vencimento para enviar aviso'),
        ]
        for chave, valor, descricao in default_configs:
            cur.execute('''
                INSERT INTO configuracoes (chave, valor, descricao)
                VALUES (%s, %s, %s)
                ON CONFLICT (chave) DO NOTHING
            ''', (chave, valor, descricao))

        conn.commit()

        # Adicionar campos de referência se não existirem (migração)
        try:
            cur.execute('ALTER TABLE clientes ADD COLUMN IF NOT EXISTS referencia TEXT')
            cur.execute('ALTER TABLE clientes ADD COLUMN IF NOT EXISTS telefone_referencia TEXT')
            cur.execute('ALTER TABLE clientes ADD COLUMN IF NOT EXISTS endereco_referencia TEXT')
            cur.execute('ALTER TABLE clientes ADD COLUMN IF NOT EXISTS chave_pix TEXT')
            cur.execute('ALTER TABLE clientes ADD COLUMN IF NOT EXISTS rg TEXT')
            cur.execute('ALTER TABLE clientes ADD COLUMN IF NOT EXISTS empresa TEXT DEFAULT \'FH1\'')
            cur.execute('ALTER TABLE usuarios ADD COLUMN IF NOT EXISTS nivel TEXT DEFAULT \'Operador\'')
            conn.commit()
        except Exception as e:
            print(f"Aviso: Erro ao adicionar campos de referência/nível: {e}")

        # Criar usuário admin padrão se não existir
        cur.execute('SELECT id FROM usuarios WHERE email = %s', ('admin@sistema.com',))
        admin_exists = cur.fetchone()
        if not admin_exists:
            admin_senha_hash = generate_password_hash('admin123')
            cur.execute(
                'INSERT INTO usuarios (nome, email, senha, tipo, nivel) VALUES (%s, %s, %s, %s, %s)',
                ('Administrador', 'admin@sistema.com', admin_senha_hash, 'admin', 'ADM')
            )
            conn.commit()
            
        # Migrar usuários existentes para o novo sistema de níveis
        cur.execute("UPDATE usuarios SET nivel='ADM' WHERE tipo='admin'")
        cur.execute("UPDATE usuarios SET nivel='Operador' WHERE tipo='operador'")
        conn.commit()

        cur.close()
        conn.close()

# --- Decoradores de Autenticação ---
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        if session.get('usuario_tipo') != 'admin':
            flash('Acesso negado. Apenas administradores podem acessar esta página.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def gerente_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'critical')
            return redirect(url_for('login'))
        
        # Buscar nível do usuário no banco de dados
        conn = get_db()
        cur = conn.cursor(row_factory=dict_row)
        cur.execute('SELECT nivel FROM usuarios WHERE id = %s', (session['usuario_id'],))
        usuario = cur.fetchone()
        cur.close()
        conn.close()
        
        if not usuario or usuario['nivel'] not in ['Gerente', 'ADM']:
            flash('Acesso não autorizado.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

def adm_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'usuario_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'warning')
            return redirect(url_for('login'))
        
        # Buscar nível do usuário no banco de dados
        conn = get_db()
        cur = conn.cursor(row_factory=dict_row)
        cur.execute('SELECT nivel FROM usuarios WHERE id = %s', (session['usuario_id'],))
        usuario = cur.fetchone()
        cur.close()
        conn.close()
        
        if not usuario or usuario['nivel'] != 'ADM':
            flash('Acesso não autorizado. Apenas administradores.', 'danger')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    return decorated_function

# --- Funções Auxiliares ---
def calcular_valor_atualizado(cobranca):
    """Calcula o valor atualizado de uma cobrança com juros e multa."""
    if cobranca['status'] != 'Pago' and cobranca['data_vencimento']:
        # Handle both string and date object formats
        if isinstance(cobranca['data_vencimento'], str):
            data_venc = datetime.strptime(cobranca['data_vencimento'], '%Y-%m-%d')
        else:
            # It's already a date object, convert to datetime
            data_venc = datetime.combine(cobranca['data_vencimento'], datetime.min.time())
        hoje = datetime.now()
        
        if hoje > data_venc:
            dias_atraso = (hoje - data_venc).days
            
            # Buscar configurações
            conn = get_db()
            cur = conn.cursor(row_factory=dict_row)
            cur.execute('SELECT chave, valor FROM configuracoes')
            config_rows = cur.fetchall()
            config = {row['chave']: row['valor'] for row in config_rows}
            cur.close()
            conn.close()
            
            dias_tolerancia = int(config.get('dias_tolerancia', 3))
            
            if dias_atraso > dias_tolerancia:
                # Aplicar multa
                taxa_multa = float(config.get('taxa_multa', 10.0))
                multa = cobranca['valor_original'] * (taxa_multa / 100)
                
                # Aplicar juros
                taxa_juros = float(config.get('taxa_juros_mensal', 2.0))
                meses_atraso = dias_atraso / 30
                juros = cobranca['valor_original'] * (taxa_juros / 100) * meses_atraso
                
                return {
                    'multa': round(multa, 2),
                    'juros': round(juros, 2),
                    'valor_total': round(cobranca['valor_original'] + multa + juros - cobranca.get('desconto', 0), 2),
                    'dias_atraso': dias_atraso
                }
    
    return {
        'multa': 0,
        'juros': 0,
        'valor_total': cobranca['valor_original'] - cobranca.get('desconto', 0),
        'dias_atraso': 0
    }

def validar_cpf_cnpj(documento):
    """Valida CPF ou CNPJ."""
    # Remove caracteres não numéricos
    documento = re.sub(r'\D', '', documento)
    
    if len(documento) == 11:  # CPF
        # Validação simplificada do CPF
        return len(documento) == 11
    elif len(documento) == 14:  # CNPJ
        # Validação simplificada do CNPJ
        return len(documento) == 14
    
    return False

# --- Rotas de Autenticação ---
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        senha = request.form['senha']
        
        conn = get_db()
        cur = conn.cursor(row_factory=dict_row)
        cur.execute('SELECT * FROM usuarios WHERE email = %s', (email,))
        usuario = cur.fetchone()
        cur.close()
        conn.close()
        
        if usuario and check_password_hash(usuario['senha'], senha):
            session['usuario_id'] = usuario['id']
            session['usuario_nome'] = usuario['nome']
            # Derive tipo from nivel
            nivel = usuario.get('nivel', 'Operador')
            tipo = 'admin' if nivel == 'ADM' else 'operador'
            session['usuario_tipo'] = tipo
            session['usuario_nivel'] = nivel
            flash(f'Bem-vindo, {usuario["nome"]}!', 'success')
            return redirect(url_for('index'))
        else:
            flash('Email ou senha incorretos.', 'danger')
    
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    flash('Você foi desconectado com sucesso.', 'info')
    return redirect(url_for('login'))

# --- Rotas do Dashboard ---
@app.route('/')
@login_required
def index():
    """Renderiza o dashboard com estatísticas e lista de cobranças."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Estatísticas gerais
    cur.execute('SELECT COUNT(*) as count FROM clientes')
    total_clientes = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM cobrancas WHERE status = 'Pendente'")
    cobrancas_pendentes = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM cobrancas WHERE status = 'Pendente' AND data_vencimento < CURRENT_DATE")
    cobrancas_vencidas = cur.fetchone()['count']
    
    # Contagem correta: Número de parcelas que foram individualmente pagas
    cur.execute("SELECT COUNT(*) as count FROM parcelas WHERE status = 'Pago'")
    parcelas_pagas = cur.fetchone()['count']
    
    # Total de parcelas para o gráfico
    cur.execute("SELECT COUNT(*) as count FROM parcelas")
    total_parcelas = cur.fetchone()['count']
    
    # Nova lógica de cálculo baseada no Saldo Devedor Total por parcela
    from datetime import date
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    
    # Buscar todas as cobranças não pagas para calcular o saldo devedor total
    cur.execute("SELECT * FROM cobrancas WHERE status = 'Pendente'")
    cobrancas_abertas = cur.fetchall()
    
    saldo_devedor_total = 0
    for cobranca in cobrancas_abertas:
        # Buscar parcelas pendentes desta cobrança
        cur.execute('''
            SELECT * FROM parcelas 
            WHERE cobranca_id = %s AND status = 'Pendente'
        ''', (cobranca['id'],))
        parcelas_pendentes = cur.fetchall()
        
        for parcela in parcelas_pendentes:
            valor_parcela_atualizado = parcela['valor'] + (parcela['multa_manual'] or 0)
            saldo_devedor_total += valor_parcela_atualizado
    
    # Calcular total recebido no mês atual
    # Usa historico_pagamentos pois é o log unificado de todos os pagamentos (avulsos e parcelas)
    cur.execute("SELECT COALESCE(SUM(valor_pago), 0) as total FROM historico_pagamentos WHERE DATE(data_pagamento) >= %s", (primeiro_dia_mes,))
    total_recebido_mes = cur.fetchone()['total']
    
    # Calcular KPIs por empresa
    empresas = ['FH1', 'FH2', 'FH3', 'FH4']
    kpis_por_empresa = {}
    
    for emp in empresas:
        # Clientes por empresa
        cur.execute('SELECT COUNT(*) as count FROM clientes WHERE empresa = %s', (emp,))
        total_clientes_emp = cur.fetchone()['count']
        
        # Saldo Devedor por empresa
        cur.execute('''
            SELECT cob.id, cob.cliente_id 
            FROM cobrancas cob
            JOIN clientes cl ON cob.cliente_id = cl.id
            WHERE cl.empresa = %s AND cob.status = 'Pendente'
        ''', (emp,))
        cobrancas_abertas_emp = cur.fetchall()
        
        saldo_devedor_emp = 0
        for cobranca_emp in cobrancas_abertas_emp:
            cur.execute('''
                SELECT * FROM parcelas 
                WHERE cobranca_id = %s AND status = 'Pendente'
            ''', (cobranca_emp['id'],))
            parcelas_pendentes_emp = cur.fetchall()
            
            for parcela_emp in parcelas_pendentes_emp:
                valor_parcela_atualizado_emp = parcela_emp['valor'] + (parcela_emp['multa_manual'] or 0)
                saldo_devedor_emp += valor_parcela_atualizado_emp
        
        kpis_por_empresa[emp] = {
            'total_clientes': total_clientes_emp,
            'saldo_devedor': saldo_devedor_emp
        }
    
    stats = {
        'total_clientes': total_clientes,
        'cobrancas_pendentes': cobrancas_pendentes,
        'cobrancas_vencidas': cobrancas_vencidas,
        'parcelas_pagas': parcelas_pagas,
        'total_parcelas': total_parcelas,
        'saldo_devedor_total': saldo_devedor_total,
        'total_recebido_mes': total_recebido_mes,
        'kpis_por_empresa': kpis_por_empresa,
    }
    
    # Cobranças recentes com informações do cliente
    cur.execute('''
        SELECT c.*, cl.nome as cliente_nome, cl.telefone, cl.email 
        FROM cobrancas c
        JOIN clientes cl ON c.cliente_id = cl.id
        WHERE c.status <> 'Pago'
        ORDER BY c.data_vencimento ASC
        LIMIT 20
    ''')
    cobrancas = cur.fetchall()
    
    # Calcular valores atualizados para cada cobrança usando nova lógica por parcela
    hoje = date.today()
    cobrancas_atualizadas = []
    for cobranca in cobrancas:
        cobranca_dict = dict(cobranca)
        saldo_devedor_calculado = 0
        
        # Buscar parcelas pendentes desta cobrança para cálculo do saldo
        cur.execute('''
            SELECT * FROM parcelas 
            WHERE cobranca_id = %s AND status = 'Pendente'
        ''', (cobranca['id'],))
        parcelas_pendentes = cur.fetchall()
        
        for parcela in parcelas_pendentes:
            valor_parcela_atualizado = parcela['valor'] + (parcela['multa_manual'] or 0)
            saldo_devedor_calculado += valor_parcela_atualizado
        
        # Adiciona atributos dinâmicos à cobrança para exibição consistente
        cobranca_dict['multa_aplicada'] = 0  # Multas agora são por parcela
        cobranca_dict['saldo_devedor_calculado'] = saldo_devedor_calculado - (cobranca_dict['valor_pago'] or 0)
        
        # Para compatibilidade com template existente
        cobranca_dict['valor_multa'] = 0  # Multas agora são por parcela
        cobranca_dict['total_a_pagar'] = cobranca_dict['saldo_devedor_calculado']
        cobranca_dict['saldo_devedor'] = cobranca_dict['saldo_devedor_calculado']
        cobranca_dict['dias_atraso'] = max([(hoje - p['data_vencimento']).days for p in parcelas_pendentes if hoje > p['data_vencimento']], default=0)
        
        cobrancas_atualizadas.append(cobranca_dict)
    
    # --- NOVA CONSULTA ---
    # Buscar clientes com parcelas vencidas e pendentes
    cur.execute('''
        SELECT DISTINCT c.*
        FROM clientes c
        INNER JOIN cobrancas cob ON cob.cliente_id = c.id
        INNER JOIN parcelas p ON p.cobranca_id = cob.id
        WHERE p.data_vencimento < %s 
        AND p.status = 'Pendente'
        ORDER BY c.nome ASC
    ''', (hoje,))
    clientes_inadimplentes = cur.fetchall()
    
    # Calcular saldo devedor para cada cliente inadimplente
    for cliente in clientes_inadimplentes:
        saldo_devedor_cliente = 0
        
        # Buscar todas as cobranças não pagas deste cliente
        cur.execute("SELECT * FROM cobrancas WHERE cliente_id = %s AND status = 'Pendente'", (cliente['id'],))
        cobrancas_abertas = cur.fetchall()
        
        for cobranca in cobrancas_abertas:
            # Buscar parcelas pendentes desta cobrança
            cur.execute('''
                SELECT * FROM parcelas 
                WHERE cobranca_id = %s AND status = 'Pendente'
            ''', (cobranca['id'],))
            parcelas_pendentes = cur.fetchall()
            
            for parcela in parcelas_pendentes:
                valor_parcela_atualizado = parcela['valor'] + (parcela['multa_manual'] or 0)
                saldo_devedor_cliente += valor_parcela_atualizado
        
        # Adiciona o valor calculado dinamicamente ao objeto do cliente
        cliente['saldo_devedor_total'] = saldo_devedor_cliente
    
    # --- FIM DA NOVA CONSULTA ---

    cur.close()
    conn.close()
    
    return render_template('index.html', 
                         stats=stats, 
                         cobrancas=cobrancas_atualizadas,
                         clientes_inadimplentes=clientes_inadimplentes,
                         usuario=session)

# --- Rotas de Clientes ---
@app.route('/clientes')
@login_required
def listar_clientes():
    """Lista clientes com filtros por status e empresa."""
    hoje = date.today()

    filtro_status = request.args.get('status', 'todos')
    filtro_empresa = request.args.get('empresa')

    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)

    base_query = '''
        SELECT 
            c.*,
            COALESCE(agg.total_cobrancas, 0) AS total_cobrancas,
            COALESCE(agg.cobrancas_pendentes, 0) AS cobrancas_pendentes,
            COALESCE(agg.valor_pendente, 0) AS valor_pendente
        FROM clientes c
        LEFT JOIN (
            SELECT 
                cliente_id,
                COUNT(id) AS total_cobrancas,
                SUM(CASE WHEN status = 'Pendente' THEN 1 ELSE 0 END) AS cobrancas_pendentes,
                SUM(CASE WHEN status = 'Pendente' THEN valor_original ELSE 0 END) AS valor_pendente
            FROM cobrancas
            GROUP BY cliente_id
        ) agg ON agg.cliente_id = c.id
    '''

    conditions = []
    params = []

    if filtro_empresa:
        conditions.append('c.empresa = %s')
        params.append(filtro_empresa)

    if filtro_status == 'atrasado':
        # Buscar clientes com parcelas vencidas e pendentes (similar ao dashboard)
        conditions.append('EXISTS (SELECT 1 FROM cobrancas cob INNER JOIN parcelas p ON p.cobranca_id = cob.id WHERE cob.cliente_id = c.id AND p.data_vencimento < %s AND p.status = %s)')
        params.append(hoje)
        params.append('Pendente')

    if conditions:
        base_query += ' WHERE ' + ' AND '.join(conditions)

    base_query += ' ORDER BY c.nome ASC'

    cur.execute(base_query, tuple(params))
    clientes = cur.fetchall()

    # Nova lógica de cálculo por parcela
    for cliente in clientes:
        saldo_devedor_cliente = 0
        cur.execute('''
            SELECT id FROM cobrancas 
            WHERE cliente_id = %s AND status = 'Pendente'
        ''', (cliente['id'],))
        cobrancas_abertas = cur.fetchall()

        for cobranca in cobrancas_abertas:
            cur.execute('''
                SELECT valor, multa_manual FROM parcelas 
                WHERE cobranca_id = %s AND status = 'Pendente'
            ''', (cobranca['id'],))
            parcelas_pendentes = cur.fetchall()

            for parcela in parcelas_pendentes:
                valor_parcela_atualizado = parcela['valor'] + (parcela['multa_manual'] or 0)
                saldo_devedor_cliente += valor_parcela_atualizado

        cliente['saldo_devedor_total'] = saldo_devedor_cliente

    cur.close()
    conn.close()

    return render_template('clientes.html', clientes=clientes, 
                           filtro_ativo=filtro_status,
                           filtro_empresa=filtro_empresa)



@app.route('/cliente/adicionar', methods=['GET', 'POST'])
@login_required
def adicionar_cliente():
    """Adiciona um novo cliente."""
    if request.method == 'POST':
        files = request.files.getlist('documentos')

        dados = {
            'nome': request.form['nome'],
            'cpf_cnpj': request.form.get('cpf_cnpj', ''),
            'rg': request.form.get('rg', ''),
            'email': request.form.get('email', ''),
            'telefone': request.form['telefone'],
            'telefone_secundario': request.form.get('telefone_secundario', ''),
            'chave_pix': request.form.get('chave_pix', ''),
            'endereco': request.form.get('endereco', ''),
            'cidade': request.form.get('cidade', ''),
            'estado': request.form.get('estado', ''),
            'cep': request.form.get('cep', ''),
            'referencia': request.form.get('referencia', ''),
            'telefone_referencia': request.form.get('telefone_referencia', ''),
            'endereco_referencia': request.form.get('endereco_referencia', ''),
            'observacoes': request.form.get('observacoes', ''),
            'empresa': request.form.get('empresa', 'FH1')
        }
        
        # Validação
        if not dados['cpf_cnpj']:
            flash('CPF/CNPJ é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['rg']:
            flash('RG é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['chave_pix']:
            flash('Chave Pix é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['referencia']:
            flash('Nome da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['telefone_referencia']:
            flash('Telefone da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['endereco_referencia']:
            flash('Endereço da referência é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['endereco']:
            flash('Endereço é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['cidade']:
            flash('Cidade é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['estado']:
            flash('Estado é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['cep']:
            flash('CEP é obrigatório.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if dados['cpf_cnpj'] and not validar_cpf_cnpj(dados['cpf_cnpj']):
            flash('CPF/CNPJ inválido.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['empresa']:
            flash('Empresa é obrigatória.', 'danger')
            return render_template('cliente_form.html', cliente=dados)
        
        conn = get_db()
        cur = conn.cursor()
        try:
            cur.execute('''
                INSERT INTO clientes (nome, cpf_cnpj, rg, email, telefone, telefone_secundario, chave_pix,
                                    endereco, cidade, estado, cep, referencia, telefone_referencia, endereco_referencia, observacoes, empresa)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', tuple(dados.values()))
            novo_cliente_id = cur.fetchone()[0]

            # Processar uploads de documentos
            if files:
                for file in files:
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        client_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(novo_cliente_id))
                        os.makedirs(client_upload_folder, exist_ok=True)
                        file_path = os.path.join(client_upload_folder, filename)
                        file.save(file_path)
                        cur.execute('INSERT INTO documentos (cliente_id, nome_ficheiro) VALUES (%s, %s)', (novo_cliente_id, filename))

            conn.commit()
            flash('Cliente adicionado com sucesso!', 'success')
            return redirect(url_for('listar_clientes'))
        except UniqueViolation:
            flash('CPF/CNPJ já cadastrado.', 'danger')
        finally:
            cur.close()
            conn.close()
    
    return render_template('cliente_form.html', cliente=None)

@app.route('/cliente/<int:cliente_id>')
@login_required
def visualizar_cliente(cliente_id):
    """Visualiza detalhes de um cliente específico."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    cur.execute('SELECT * FROM clientes WHERE id = %s', (cliente_id,))
    cliente = cur.fetchone()
    if not cliente:
        flash('Cliente não encontrado.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('listar_clientes'))
    
    # Cobranças do cliente
    cur.execute('''
        SELECT * FROM cobrancas 
        WHERE cliente_id = %s 
        ORDER BY data_vencimento DESC
    ''', (cliente_id,))
    cobrancas = cur.fetchall()
    
    # Histórico de pagamentos
    cur.execute('''
        SELECT h.*, c.descricao as cobranca_descricao
        FROM historico_pagamentos h
        JOIN cobrancas c ON h.cobranca_id = c.id
        WHERE h.cliente_id = %s
        ORDER BY h.data_pagamento DESC
    ''', (cliente_id,))
    historico = cur.fetchall()
    
    # Carregar documentos do cliente
    cur = get_db().cursor(row_factory=dict_row)
    cur.execute('SELECT * FROM documentos WHERE cliente_id = %s ORDER BY id DESC', (cliente_id,))
    documentos = cur.fetchall()
    cur.close()

    # Nova lógica de cálculo por parcela
    hoje = date.today()
    cobrancas_processadas = []
    if cobrancas:  # Garante que só vai iterar se houver cobranças
        for cobranca in cobrancas:
            # Converter dict para um objeto mutável
            cobranca_dict = dict(cobranca)
            saldo_devedor_cobranca = Decimal('0.00')
            multa_total_cobranca = 0
            cobranca_dict['parcelas_com_multa'] = []  # Lista para guardar dados das parcelas para o template

            # Buscar parcelas para esta cobrança
            conn_temp = get_db()
            cur_temp = conn_temp.cursor(row_factory=dict_row)
            cur_temp.execute('''
                SELECT * FROM parcelas 
                WHERE cobranca_id = %s 
                ORDER BY numero_parcela ASC
            ''', (cobranca['id'],))
            parcelas = cur_temp.fetchall()
            cur_temp.close()
            conn_temp.close()

            for parcela in parcelas:
                parcela_dict = dict(parcela)
                multa_manual = Decimal(str(parcela_dict.get('multa_manual') or 0))
                valor_original = Decimal(str(parcela_dict['valor']))
                valor_pago_atual = Decimal(str(parcela_dict.get('valor_pago') or 0))
                valor_atualizado = valor_original + multa_manual
                valor_restante = max(valor_atualizado - valor_pago_atual, Decimal('0.00'))
                parcela_dict['multa_manual'] = float(multa_manual)
                parcela_dict['valor_atualizado'] = valor_atualizado

                if parcela_dict['status'] == 'Pendente':
                    saldo_devedor_cobranca += valor_restante

                # Guarda os dados calculados para exibir no template
                cobranca_dict['parcelas_com_multa'].append({
                    'numero': parcela_dict['numero_parcela'],
                    'vencimento': parcela_dict['data_vencimento'],
                    'valor_original': parcela_dict['valor'],
                    'multa_manual': parcela_dict['multa_manual'],
                    'valor_atualizado': parcela_dict['valor_atualizado'],
                    'status': parcela_dict['status'],
                    'id': parcela_dict['id'],
                    'valor_pago_atual': float(valor_pago_atual),
                    'valor_restante': float(valor_restante)
                })

            # Adiciona atributos dinâmicos à cobrança para exibição consistente
            cobranca_dict['multa_aplicada'] = 0  # Multas agora são por parcela
            cobranca_dict['saldo_devedor_calculado'] = float(saldo_devedor_cobranca)
            
            # Para compatibilidade com template existente
            cobranca_dict['valor_multa'] = 0  # Multas agora são por parcela
            cobranca_dict['total_a_pagar'] = cobranca_dict['saldo_devedor_calculado']
            cobranca_dict['saldo_devedor'] = cobranca_dict['saldo_devedor_calculado']
            
            cobrancas_processadas.append(cobranca_dict)
    
    return render_template('cliente_detalhes.html', 
                         cliente=cliente, 
                         cobrancas=cobrancas_processadas,
                         historico=historico,
                         documentos=documentos,
                         today=hoje)

@app.route('/cliente/<int:cliente_id>/editar', methods=['GET', 'POST'])
@login_required
def editar_cliente(cliente_id):
    """Edita um cliente existente."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    cur.execute('SELECT * FROM clientes WHERE id = %s', (cliente_id,))
    cliente = cur.fetchone()
    if not cliente:
        flash('Cliente não encontrado.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('listar_clientes'))
    
    if request.method == 'POST':
        dados = {
            'nome': request.form['nome'],
            'cpf_cnpj': request.form.get('cpf_cnpj', ''),
            'rg': request.form.get('rg', ''),
            'email': request.form.get('email', ''),
            'telefone': request.form['telefone'],
            'telefone_secundario': request.form.get('telefone_secundario', ''),
            'chave_pix': request.form.get('chave_pix', ''),
            'endereco': request.form.get('endereco', ''),
            'cidade': request.form.get('cidade', ''),
            'estado': request.form.get('estado', ''),
            'cep': request.form.get('cep', ''),
            'referencia': request.form.get('referencia', ''),
            'telefone_referencia': request.form.get('telefone_referencia', ''),
            'endereco_referencia': request.form.get('endereco_referencia', ''),
            'observacoes': request.form.get('observacoes', ''),
            'empresa': request.form.get('empresa', 'FH1')
        }
        
        # Validação
        if not dados['cpf_cnpj']:
            flash('CPF/CNPJ é obrigatório.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['rg']:
            flash('RG é obrigatório.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['chave_pix']:
            flash('Chave Pix é obrigatória.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['referencia']:
            flash('Nome da referência é obrigatório.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['telefone_referencia']:
            flash('Telefone da referência é obrigatório.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['endereco_referencia']:
            flash('Endereço da referência é obrigatório.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['endereco']:
            flash('Endereço é obrigatório.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['cidade']:
            flash('Cidade é obrigatória.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['estado']:
            flash('Estado é obrigatório.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['cep']:
            flash('CEP é obrigatório.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if dados['cpf_cnpj'] and not validar_cpf_cnpj(dados['cpf_cnpj']):
            flash('CPF/CNPJ inválido.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        if not dados['empresa']:
            flash('Empresa é obrigatória.', 'danger')
            cur.close()
            conn.close()
            return render_template('cliente_form.html', cliente=dados)
        
        try:
            cur.execute('''
                UPDATE clientes 
                SET nome=%s, cpf_cnpj=%s, rg=%s, email=%s, telefone=%s, telefone_secundario=%s, chave_pix=%s,
                    endereco=%s, cidade=%s, estado=%s, cep=%s, referencia=%s, telefone_referencia=%s, endereco_referencia=%s, observacoes=%s, empresa=%s,
                    atualizado_em=CURRENT_TIMESTAMP
                WHERE id=%s
            ''', (dados['nome'], dados['cpf_cnpj'], dados['rg'], dados['email'], dados['telefone'], 
                 dados['telefone_secundario'], dados['chave_pix'], dados['endereco'], dados['cidade'], 
                 dados['estado'], dados['cep'], dados['referencia'], dados['telefone_referencia'], dados['endereco_referencia'], dados['observacoes'], dados['empresa'], cliente_id))

            # Processar uploads de documentos (se houver)
            files = request.files.getlist('documentos')
            if files:
                for file in files:
                    if file and file.filename:
                        filename = secure_filename(file.filename)
                        client_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], str(cliente_id))
                        os.makedirs(client_upload_folder, exist_ok=True)
                        file_path = os.path.join(client_upload_folder, filename)
                        file.save(file_path)
                        cur.execute('INSERT INTO documentos (cliente_id, nome_ficheiro) VALUES (%s, %s)', (cliente_id, filename))

            conn.commit()
            flash('Cliente atualizado com sucesso!', 'success')
            cur.close()
            conn.close()
            return redirect(url_for('listar_clientes'))
        except UniqueViolation:
            flash('CPF/CNPJ já cadastrado.', 'danger')
        finally:
            cur.close()
            conn.close()
    
    cur.close()
    conn.close()
    return render_template('cliente_form.html', cliente=cliente)

@app.route('/cliente/<int:cliente_id>/deletar', methods=['POST'])
@login_required
def deletar_cliente(cliente_id):
    """Deleta um cliente e todas suas cobranças relacionadas."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Verificar se o cliente existe
    cur.execute('SELECT nome FROM clientes WHERE id = %s', (cliente_id,))
    cliente = cur.fetchone()
    if not cliente:
        flash('Cliente não encontrado.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('listar_clientes'))
    
    try:
        # Deletar cliente (cascata deletará cobranças e histórico)
        cur.execute('DELETE FROM clientes WHERE id = %s', (cliente_id,))
        conn.commit()
        flash(f'Cliente "{cliente["nome"]}" e todas as cobranças relacionadas foram excluídos com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao excluir cliente: {str(e)}', 'danger')
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('listar_clientes'))

@app.route('/cobranca/adicionar', methods=['GET', 'POST'])
@login_required
def adicionar_cobranca():
    """Adiciona uma nova cobrança."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    cur.execute('SELECT id, nome FROM clientes ORDER BY nome')
    clientes = cur.fetchall()
    cur.close()
    conn.close()
    
    if request.method == 'POST':
        # Obter dados do formulário
        cliente_id = int(request.form['cliente_id'])
        descricao = request.form['descricao']
        valor_emprestimo = float(request.form['valor_emprestimo'])
        taxa_juros = float(request.form['taxa_juros'])
        data_primeira_parcela = request.form['data_vencimento']
        
        # Validar se a data não é domingo
        data_venc = datetime.strptime(data_primeira_parcela, '%Y-%m-%d')
        if data_venc.weekday() == 6:  # 6 = domingo (0=segunda, 6=domingo)
            flash('Domingos não são permitidos para data de vencimento. Selecione outro dia.', 'danger')
            return render_template('cobranca_form.html', clientes=clientes, cobranca=None)
        
        # Determinar número de parcelas baseado na taxa de juros
        if taxa_juros == 30:
            numero_parcelas = 10
        elif taxa_juros == 60:
            numero_parcelas = 15
        else:
            flash('Taxa de juros inválida. Use 30% ou 60%.', 'danger')
            return render_template('cobranca_form.html', clientes=clientes, cobranca=None)
        
        # Calcular valor total com juros
        valor_devido_total = valor_emprestimo * (1 + (taxa_juros / 100))
        valor_parcela = valor_devido_total / numero_parcelas
        
        conn = get_db()
        cur = conn.cursor()
        
        try:
            # Criar a cobrança principal
            cur.execute('''
                INSERT INTO cobrancas (cliente_id, descricao, valor_original, taxa_juros, valor_total, data_vencimento, 
                                     tipo_cobranca, numero_parcelas)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (cliente_id, descricao, valor_emprestimo, taxa_juros, valor_devido_total, 
                 data_primeira_parcela, 'Parcelada', numero_parcelas))
            
            nova_cobranca_id = cur.fetchone()[0]
            
            # Gerar as parcelas diárias
            data_vencimento_atual = data_venc.date()
            
            for i in range(numero_parcelas):
                # Pular domingos
                while data_vencimento_atual.weekday() == 6:  # 6 = domingo
                    data_vencimento_atual += timedelta(days=1)
                
                # Criar a parcela
                cur.execute('''
                    INSERT INTO parcelas (cobranca_id, numero_parcela, valor, data_vencimento, status)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (nova_cobranca_id, i + 1, valor_parcela, data_vencimento_atual, 'Pendente'))
                
                # Incrementar para a próxima parcela (próximo dia)
                data_vencimento_atual += timedelta(days=1)
            
            conn.commit()
            flash(f'Cobrança criada com sucesso! {numero_parcelas} parcelas diárias foram geradas.', 'success')
            
        except Exception as e:
            conn.rollback()
            flash(f'Erro ao criar cobrança: {str(e)}', 'danger')
        finally:
            cur.close()
            conn.close()
        
        return redirect(url_for('index'))
    
    return render_template('cobranca_form.html', clientes=clientes, cobranca=None)

@app.route('/cobranca/<int:cobranca_id>/cancelar', methods=['POST'])
@login_required
def cancelar_cobranca(cobranca_id):
    """DELETA uma cobrança e todos os seus registros associados (parcelas, pagamentos)."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row) # Usar dict_row para facilitar
    
    # Verificar se a cobrança existe e pegar o cliente_id para redirecionar
    cur.execute('SELECT id, cliente_id FROM cobrancas WHERE id = %s', (cobranca_id,))
    cobranca = cur.fetchone()
    
    if not cobranca:
        flash('Cobrança não encontrada.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('index'))
    
    # Armazena o ID do cliente para o redirecionamento
    cliente_id_redirect = cobranca['cliente_id']
    
    try:
        # Deletar a cobrança. O 'ON DELETE CASCADE' cuidará das parcelas e pagamentos.
        cur.execute('DELETE FROM cobrancas WHERE id = %s', (cobranca_id,))
        
        conn.commit()
        flash('Cobrança e todos os seus dados foram DELETADOS com sucesso!', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao deletar cobrança: {str(e)}', 'danger')
    finally:
        cur.close()
        conn.close()
    
    # Redireciona de volta para a página do cliente
    return redirect(url_for('visualizar_cliente', cliente_id=cliente_id_redirect))

@app.route('/cobranca/<int:id>/registrar_pagamento', methods=['POST'])
@login_required
def registrar_pagamento(id):
    """Registra um pagamento genérico para uma cobrança."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Buscar cobrança
    cur.execute('SELECT * FROM cobrancas WHERE id = %s', (id,))
    cobranca = cur.fetchone()
    if not cobranca:
        flash('Cobrança não encontrada.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('index'))
    
    # Obter dados do formulário
    valor_a_pagar = float(request.form['valor_pago'])
    observacao_pagamento = request.form.get('observacao_pagamento', '')
    
    if valor_a_pagar <= 0:
        flash('O valor do pagamento deve ser maior que zero.', 'warning')
        cur.close()
        conn.close()
        return redirect(url_for('visualizar_cliente', cliente_id=cobranca['cliente_id']))
    
    # Criar um novo registro de pagamento
    cur.execute('''
        INSERT INTO pagamentos (cobranca_id, valor_pago, observacao, forma_pagamento, usuario_id)
        VALUES (%s, %s, %s, %s, %s)
    ''', (id, valor_a_pagar, observacao_pagamento, 'Dinheiro', session.get('usuario_id')))
    
    # Atualizar o total pago na cobrança
    novo_valor_pago = (cobranca['valor_pago'] or 0) + valor_a_pagar
    cur.execute('''
        UPDATE cobrancas 
        SET valor_pago=%s, atualizado_em=CURRENT_TIMESTAMP
        WHERE id=%s
    ''', (novo_valor_pago, id))
    
    # Registrar também no histórico de pagamentos para compatibilidade
    cur.execute('''
        INSERT INTO historico_pagamentos (cobranca_id, cliente_id, valor_pago, forma_pagamento, observacoes, usuario_id)
        VALUES (%s, %s, %s, %s, %s, %s)
    ''', (id, cobranca['cliente_id'], valor_a_pagar, 'Dinheiro', observacao_pagamento, session.get('usuario_id')))
    
    conn.commit()
    cur.close()
    conn.close()
    flash(f'Pagamento de R$ {valor_a_pagar:.2f} registrado com sucesso.', 'info')
    return redirect(url_for('visualizar_cliente', cliente_id=cobranca['cliente_id']))


@app.route('/cobranca/<int:cobranca_id>/pagamentos')
@login_required
def visualizar_pagamentos_cobranca(cobranca_id):
    """Visualiza o histórico de pagamentos de uma cobrança específica."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Buscar cobrança
    cur.execute('''
        SELECT c.*, cl.nome as cliente_nome 
        FROM cobrancas c
        JOIN clientes cl ON c.cliente_id = cl.id
        WHERE c.id = %s
    ''', (cobranca_id,))
    cobranca = cur.fetchone()
    
    if not cobranca:
        flash('Cobrança não encontrada.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('index'))
    
    # Buscar pagamentos individuais
    cur.execute('''
        SELECT p.*, u.nome as usuario_nome
        FROM pagamentos p
        LEFT JOIN usuarios u ON p.usuario_id = u.id
        WHERE p.cobranca_id = %s
        ORDER BY p.data_pagamento DESC, p.criado_em DESC
    ''', (cobranca_id,))
    pagamentos = cur.fetchall()
    
    cur.close()
    conn.close()
    
    return render_template('pagamentos_cobranca.html', 
                         cobranca=cobranca, 
                         pagamentos=pagamentos)

@app.route('/parcela/<int:id>/pagar', methods=['POST'])
@login_required
def marcar_parcela_paga(id):
    """Marca uma parcela como paga e atualiza o valor_pago da cobrança."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Buscar parcela e cobrança relacionada
    cur.execute('''
        SELECT p.*, c.cliente_id, c.valor_total
        FROM parcelas p
        JOIN cobrancas c ON p.cobranca_id = c.id
        WHERE p.id = %s
    ''', (id,))
    parcela = cur.fetchone()
    
    if not parcela:
        flash('Parcela não encontrada.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('index'))
    
    # Recálculo robusto do valor devido: valor original + multa_manual (trata None como 0)
    valor_original = Decimal(str(parcela['valor']))
    multa_manual = Decimal(str(parcela.get('multa_manual') or 0))
    valor_devido = valor_original + multa_manual
    
    # Soma correta do valor pago: valor_pago existente (trata None como 0) + valor_recebido_agora
    valor_pago_existente = Decimal(str(parcela.get('valor_pago') or 0))
    saldo_restante = max(valor_devido - valor_pago_existente, Decimal('0.00'))

    if parcela['status'] == 'Pago' or saldo_restante <= 0:
        flash('Esta parcela já foi paga.', 'info')
        cur.close()
        conn.close()
        return redirect(url_for('visualizar_cliente', cliente_id=parcela['cliente_id']))

    # Obter valor recebido agora do formulário
    valor_input = request.form.get('valor_pago')
    valor_recebido_agora = None
    if valor_input:
        valor_normalizado = str(valor_input).replace(',', '.').strip()
        if valor_normalizado:
            try:
                valor_recebido_agora = Decimal(valor_normalizado)
            except InvalidOperation:
                cur.close()
                conn.close()
                flash('Valor de pagamento inválido.', 'danger')
                return redirect(url_for('visualizar_cliente', cliente_id=parcela['cliente_id']))
            if valor_recebido_agora <= 0:
                valor_recebido_agora = None

    # Se não foi informado valor, usar o saldo restante completo
    if not valor_recebido_agora:
        valor_recebido_agora = saldo_restante

    # Limitar o valor recebido ao saldo restante disponível
    if saldo_restante > 0 and valor_recebido_agora > saldo_restante:
        valor_recebido_agora = saldo_restante

    valor_recebido_agora = valor_recebido_agora.quantize(Decimal('0.01'))

    if valor_recebido_agora <= 0:
        flash('Não há saldo pendente para esta parcela.', 'info')
        cur.close()
        conn.close()
        return redirect(url_for('visualizar_cliente', cliente_id=parcela['cliente_id']))

    # Calcular total pago após esta transação (cumulativo)
    total_pago = (valor_pago_existente + valor_recebido_agora).quantize(Decimal('0.01'))
    
    # Comparação com tolerância para evitar erros de ponto flutuante (FIX PRINCIPAL)
    tolerancia = Decimal('0.01')
    if total_pago >= (valor_devido - tolerancia):
        status_novo = 'Pago'
    else:
        status_novo = 'Pendente'
    
    data_pagamento_valor = date.today() if status_novo == 'Pago' else parcela.get('data_pagamento')
    
    try:
        # Atualizar parcela: valor_pago cumulativo e status calculado
        cur.execute('''
            UPDATE parcelas 
            SET status=%s,
                valor_pago=%s,
                forma_pagamento='Dinheiro',
                data_pagamento=%s,
                atualizado_em=CURRENT_TIMESTAMP
            WHERE id=%s
        ''', (status_novo, total_pago, data_pagamento_valor, id))
        
        # Atualizar valor pago na cobrança principal (incrementa com o valor desta transação)
        cur.execute('''
            UPDATE cobrancas 
            SET valor_pago = COALESCE(valor_pago, 0) + %s, atualizado_em=CURRENT_TIMESTAMP
            WHERE id=%s
        ''', (valor_recebido_agora, parcela['cobranca_id']))
        
        # Registrar no histórico de pagamentos apenas com o valor pago NESTA transação (não cumulativo)
        cur.execute('''
            INSERT INTO historico_pagamentos (cobranca_id, cliente_id, valor_pago, forma_pagamento, observacoes, usuario_id)
            VALUES (%s, %s, %s, %s, %s, %s)
        ''', (
            parcela['cobranca_id'],
            parcela['cliente_id'],
            valor_recebido_agora,
            'Dinheiro',
            f'Pagamento da parcela {parcela["numero_parcela"]}',
            session.get('usuario_id')
        ))
        
        # Verificar se todas as parcelas da cobrança foram pagas
        cur.execute('SELECT COUNT(*) as total, SUM(CASE WHEN status = \'Pago\' THEN 1 ELSE 0 END) as pagas FROM parcelas WHERE cobranca_id = %s', (parcela['cobranca_id'],))
        status_parcelas = cur.fetchone()
        
        if status_parcelas['pagas'] == status_parcelas['total']:
            # Todas as parcelas foram pagas, atualizar cobrança principal
            cur.execute('''
                UPDATE cobrancas 
                SET status='Pago', data_pagamento=CURRENT_DATE, atualizado_em=CURRENT_TIMESTAMP
                WHERE id=%s
            ''', (parcela['cobranca_id'],))
            flash(f'Parcela {parcela["numero_parcela"]} paga. Todas as parcelas foram pagas e a cobrança foi liquidada!', 'success')
        else:
            if status_novo == 'Pago':
                flash(f'Parcela {parcela["numero_parcela"]} quitada com pagamento de R$ {float(valor_recebido_agora):.2f}.', 'success')
            else:
                saldo_pos_pagamento = max(float((valor_devido - total_pago).quantize(Decimal('0.01'))), 0.0)
                flash(f'Pagamento parcial de R$ {float(valor_recebido_agora):.2f} registrado para a parcela {parcela["numero_parcela"]}. Saldo restante: R$ {saldo_pos_pagamento:.2f}.', 'info')
        
        conn.commit()
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao processar pagamento: {str(e)}', 'danger')
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('visualizar_cliente', cliente_id=parcela['cliente_id']))

@app.route('/parcela/<int:id>/editar_multa', methods=['POST'])
@login_required
def editar_multa_parcela(id):
    """Edita a multa manual de uma parcela."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Buscar parcela e cobrança relacionada
    cur.execute('''
        SELECT p.*, c.cliente_id
        FROM parcelas p
        JOIN cobrancas c ON p.cobranca_id = c.id
        WHERE p.id = %s
    ''', (id,))
    parcela = cur.fetchone()
    
    if not parcela:
        flash('Parcela não encontrada.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('index'))
    
    multa_input = request.form.get('multa_manual_parcela')

    try:
        if multa_input and multa_input.strip() != '':
            multa_valor = float(multa_input)
            if multa_valor < 0:
                flash('O valor da multa manual não pode ser negativo.', 'danger')
            else:
                cur.execute('''
                    UPDATE parcelas 
                    SET multa_manual=%s, atualizado_em=CURRENT_TIMESTAMP
                    WHERE id=%s
                ''', (multa_valor, id))
                flash(f'Multa manual de R$ {multa_valor:.2f} definida para a parcela {parcela["numero_parcela"]}.', 'success')
        else: 
            cur.execute('''
                UPDATE parcelas 
                SET multa_manual=NULL, atualizado_em=CURRENT_TIMESTAMP
                WHERE id=%s
            ''', (id,))
            flash(f'Multa manual removida da parcela {parcela["numero_parcela"]}.', 'info')

        conn.commit()
    except ValueError:
        flash('Valor da multa inválido. Por favor, insira um número.', 'danger')
    except Exception as e:
        flash(f'Erro ao atualizar multa: {str(e)}', 'danger')
    finally:
        cur.close()
        conn.close()

    return redirect(url_for('visualizar_cliente', cliente_id=parcela['cliente_id']))

@app.route('/parcela/<int:id>/editar_data', methods=['POST'])
@login_required
def editar_data_parcela(id):
    """Edita a data de vencimento de uma parcela."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Buscar parcela e cobrança relacionada
    cur.execute('''
        SELECT p.*, c.cliente_id
        FROM parcelas p
        JOIN cobrancas c ON p.cobranca_id = c.id
        WHERE p.id = %s
    ''', (id,))
    parcela = cur.fetchone()
    
    if not parcela:
        flash('Parcela não encontrada.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('index'))
    
    nova_data_str = request.form.get('nova_data_vencimento')
    
    if not nova_data_str:
        flash('Data de vencimento inválida.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('visualizar_cliente', cliente_id=parcela['cliente_id']))
    
    try:
        nova_data = datetime.strptime(nova_data_str, '%Y-%m-%d').date()
        
        # Validação para não permitir data no domingo (opcional, mas recomendado)
        if nova_data.weekday() == 6:  # 6 = Domingo
            flash('A data de vencimento não pode ser num domingo.', 'warning')
            cur.close()
            conn.close()
            return redirect(url_for('visualizar_cliente', cliente_id=parcela['cliente_id']))
        
        cur.execute('''
            UPDATE parcelas 
            SET data_vencimento=%s, atualizado_em=CURRENT_TIMESTAMP
            WHERE id=%s
        ''', (nova_data, id))
        
        conn.commit()
        flash(f'Data de vencimento da parcela {parcela["numero_parcela"]} atualizada para {nova_data.strftime("%d/%m/%Y")}.', 'success')
    except ValueError:
        flash('Formato de data inválido.', 'danger')
    except Exception as e:
        flash(f'Erro ao atualizar data: {str(e)}', 'danger')
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('visualizar_cliente', cliente_id=parcela['cliente_id']))

@app.route('/cobrancas/editar/<int:id>', methods=['GET', 'POST'])
@login_required
def editar_cobranca(id):
    """Edita uma cobrança existente."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Buscar cobrança
    cur.execute('SELECT * FROM cobrancas WHERE id = %s', (id,))
    cobranca = cur.fetchone()
    if not cobranca:
        flash('Cobrança não encontrada.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('index'))
    
    # Buscar cliente da cobrança
    cur.execute('SELECT * FROM clientes WHERE id = %s', (cobranca['cliente_id'],))
    cliente = cur.fetchone()
    if not cliente:
        flash('Cliente não encontrado.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            # Obter dados do formulário
            valor_emprestimo_str = request.form.get('valor_emprestimo')
            if not valor_emprestimo_str:
                flash('Valor emprestado é obrigatório.', 'danger')
                cur.close()
                conn.close()
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            novo_valor_emprestimo = float(valor_emprestimo_str)
            if novo_valor_emprestimo <= 0:
                flash('Valor emprestado deve ser maior que zero.', 'danger')
                cur.close()
                conn.close()
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            nova_data_vencimento_str = request.form.get('data_vencimento')
            if not nova_data_vencimento_str:
                flash('Data de vencimento é obrigatória.', 'danger')
                cur.close()
                conn.close()
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            # Validar formato da data (YYYY-MM-DD)
            nova_data_vencimento = datetime.strptime(nova_data_vencimento_str, '%Y-%m-%d').date()
            
            # Validar se a data não é domingo
            if nova_data_vencimento.weekday() == 6:  # 6 = Domingo
                flash('Domingos não são permitidos para data de vencimento. Selecione outro dia.', 'danger')
                cur.close()
                conn.close()
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            # Obter taxa de juros do formulário (ou usar a existente se não fornecida)
            taxa_juros_str = request.form.get('taxa_juros')
            if taxa_juros_str:
                taxa_juros = float(taxa_juros_str)
            else:
                taxa_juros = cobranca['taxa_juros']
            
            if not taxa_juros:
                flash('Taxa de juros é obrigatória.', 'danger')
                cur.close()
                conn.close()
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            # Determinar número de parcelas baseado na taxa de juros
            if taxa_juros == 30:
                numero_parcelas = 10
            elif taxa_juros == 60:
                numero_parcelas = 15
            else:
                flash('Taxa de juros inválida. Use 30% ou 60%.', 'danger')
                cur.close()
                conn.close()
                return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)
            
            # Calcular valor total com juros
            valor_total = novo_valor_emprestimo * (1 + (taxa_juros / 100))
            valor_parcela = valor_total / numero_parcelas
            
            # Apagar todas as parcelas antigas associadas a esta cobrança
            cur.execute('DELETE FROM parcelas WHERE cobranca_id = %s', (id,))
            
            # Apagar histórico de pagamentos da cobrança (já que as parcelas foram apagadas)
            cur.execute('DELETE FROM historico_pagamentos WHERE cobranca_id = %s', (id,))
            
            # Re-gerar as parcelas diárias com a nova data inicial
            data_vencimento_atual = nova_data_vencimento
            
            for i in range(numero_parcelas):
                # Pular domingos
                while data_vencimento_atual.weekday() == 6:  # 6 = domingo
                    data_vencimento_atual += timedelta(days=1)
                
                # Criar a parcela
                cur.execute('''
                    INSERT INTO parcelas (cobranca_id, numero_parcela, valor, data_vencimento, status)
                    VALUES (%s, %s, %s, %s, %s)
                ''', (id, i + 1, valor_parcela, data_vencimento_atual, 'Pendente'))
                
                # Incrementar para a próxima parcela (próximo dia)
                data_vencimento_atual += timedelta(days=1)
            
            # Atualizar a cobrança principal
            # Obter a data da última parcela para atualizar data_vencimento da cobrança
            ultima_data_vencimento = data_vencimento_atual - timedelta(days=1)
            
            cur.execute('''
                UPDATE cobrancas 
                SET valor_original=%s, valor_total=%s, taxa_juros=%s, data_vencimento=%s, 
                    numero_parcelas=%s, valor_pago=0, status='Pendente', atualizado_em=CURRENT_TIMESTAMP
                WHERE id=%s
            ''', (novo_valor_emprestimo, valor_total, taxa_juros, nova_data_vencimento, numero_parcelas, id))
            
            # Confirmar transação
            conn.commit()
            
            flash(f'Cobrança atualizada com sucesso! {numero_parcelas} parcelas foram re-geradas a partir da nova data.', 'success')
            
        except ValueError as e:
            conn.rollback()
            flash(f'Erro ao processar dados: {str(e)}', 'danger')
        except Exception as e:
            conn.rollback()
            flash(f'Erro ao atualizar cobrança: {str(e)}', 'danger')
        finally:
            cur.close()
            conn.close()
        
        return redirect(url_for('visualizar_cliente', cliente_id=cobranca['cliente_id']))
    
    cur.close()
    conn.close()
    return render_template('cobranca_form.html', cobranca=cobranca, cliente=cliente)












# --- Rotas de Usuários ---
@app.route('/usuarios')
@login_required
@adm_required
def listar_usuarios():
    """Lista todos os usuários do sistema."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    cur.execute('SELECT * FROM usuarios ORDER BY nome')
    usuarios = cur.fetchall()
    cur.close()
    conn.close()
    
    return render_template('usuarios.html', usuarios=usuarios)

@app.route('/usuario/adicionar', methods=['GET', 'POST'])
@login_required
@adm_required
def adicionar_usuario():
    """Adiciona um novo usuário."""
    if request.method == 'POST':
        dados = {
            'nome': request.form['nome'],
            'email': request.form['email'],
            'senha': request.form['senha'],
            'nivel': request.form.get('nivel', 'Operador')
        }
        
        # Automatically derive tipo from nivel
        dados['tipo'] = 'admin' if dados['nivel'] == 'ADM' else 'operador'
        
        conn = get_db()
        cur = conn.cursor()
        try:
            senha_hash = generate_password_hash(dados['senha'])
            cur.execute(
                'INSERT INTO usuarios (nome, email, senha, tipo, nivel) VALUES (%s, %s, %s, %s, %s)',
                (dados['nome'], dados['email'], senha_hash, dados['tipo'], dados['nivel'])
            )
            conn.commit()
            flash('Usuário adicionado com sucesso!', 'success')
            return redirect(url_for('listar_usuarios'))
        except UniqueViolation:
            flash('Email já cadastrado.', 'danger')
        finally:
            cur.close()
            conn.close()
    
    return render_template('usuario_form.html', usuario=None)

@app.route('/usuario/<int:usuario_id>', methods=['GET', 'POST'])
@login_required
@adm_required
def editar_usuario(usuario_id):
    """Edita um usuário existente."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    cur.execute('SELECT * FROM usuarios WHERE id = %s', (usuario_id,))
    usuario = cur.fetchone()
    if not usuario:
        flash('Usuário não encontrado.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('listar_usuarios'))
    
    if request.method == 'POST':
        dados = {
            'nome': request.form['nome'],
            'email': request.form['email'],
            'nivel': request.form.get('nivel', 'Operador')
        }
        
        # Automatically derive tipo from nivel
        dados['tipo'] = 'admin' if dados['nivel'] == 'ADM' else 'operador'
        
        # Se uma nova senha foi fornecida
        if request.form.get('senha'):
            dados['senha'] = generate_password_hash(request.form['senha'])
            cur.execute(
                'UPDATE usuarios SET nome=%s, email=%s, senha=%s, tipo=%s, nivel=%s WHERE id=%s',
                (dados['nome'], dados['email'], dados['senha'], dados['tipo'], dados['nivel'], usuario_id)
            )
        else:
            cur.execute(
                'UPDATE usuarios SET nome=%s, email=%s, tipo=%s, nivel=%s WHERE id=%s',
                (dados['nome'], dados['email'], dados['tipo'], dados['nivel'], usuario_id)
            )
        
        conn.commit()
        cur.close()
        conn.close()
        flash('Usuário atualizado com sucesso!', 'success')
        return redirect(url_for('listar_usuarios'))
    
    cur.close()
    conn.close()
    return render_template('usuario_form.html', usuario=usuario)

@app.route('/usuario/<int:usuario_id>/deletar', methods=['POST'])
@login_required
@adm_required
def excluir_usuario(usuario_id):
    """Exclui um usuário do sistema."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Verificar se o usuário existe
    cur.execute('SELECT nome FROM usuarios WHERE id = %s', (usuario_id,))
    usuario = cur.fetchone()
    if not usuario:
        flash('Usuário não encontrado.', 'danger')
        cur.close()
        conn.close()
        return redirect(url_for('listar_usuarios'))
    
    try:
        # Excluir usuário
        cur.execute('DELETE FROM usuarios WHERE id = %s', (usuario_id,))
        conn.commit()
        flash(f'Usuário "{usuario["nome"]}" foi excluído com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao excluir usuário: {str(e)}', 'danger')
    finally:
        cur.close()
        conn.close()
    
    return redirect(url_for('listar_usuarios'))

# --- Rotas do Calendário ---
@app.route('/calendario')
@login_required
def calendario():
    """Renderiza a página do calendário."""
    return render_template('calendario.html')

@app.route('/api/eventos')
@login_required
def api_eventos():
    """API para buscar eventos do calendário."""
    # Filtra apenas cobranças com data de vencimento não nula
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    cur.execute('''
        SELECT c.*, cl.nome as cliente_nome, cl.telefone as cliente_telefone, cl.email as cliente_email
        FROM cobrancas c
        JOIN clientes cl ON c.cliente_id = cl.id
        WHERE c.data_vencimento IS NOT NULL
        ORDER BY c.data_vencimento ASC
    ''')
    cobrancas = cur.fetchall()
    cur.close()
    conn.close()
    
    eventos = []
    for cobranca in cobrancas:
        # Garante que o cliente existe para evitar erros
        if cobranca['cliente_nome']:
            eventos.append({
                'title': f"{cobranca['cliente_nome']} - R$ {cobranca['valor_original']:.2f}",
                'start': cobranca['data_vencimento'].isoformat(),
                'url': url_for('visualizar_cliente', cliente_id=cobranca['cliente_id'])
            })
    
    return jsonify(eventos)

# --- Rota para servir uploads ---
@app.route('/uploads/<int:cliente_id>/<filename>')
@login_required
def uploaded_file(cliente_id, filename):
    # --- CORREÇÃO ---
    # 1. Obter o caminho absoluto para a pasta 'uploads' usando app.root_path
    # app.root_path é o caminho absoluto para o diretório onde app.py está
    uploads_dir = os.path.join(app.root_path, app.config['UPLOAD_FOLDER'])
    
    # 2. Construir o caminho para a pasta específica do cliente
    client_dir = os.path.join(uploads_dir, str(cliente_id))
    
    # 3. Construir o caminho completo para o arquivo
    file_path = os.path.join(client_dir, filename)
    
    # DEBUG: Imprime o caminho completo
    print(f"--- DEBUG: A tentar servir o ficheiro: {file_path} ---")
    
    # 4. Verificar se o arquivo existe
    if not os.path.isfile(file_path):
        print(f"--- DEBUG ERROR: Ficheiro não encontrado: {file_path} ---")
        abort(404)
    
    # 5. Tenta servir o ficheiro
    try:
        # Passamos o diretório absoluto do cliente e o nome do arquivo
        print(f"--- DEBUG SUCCESS: A servir {filename} de {client_dir} ---")
        return send_from_directory(client_dir, filename, as_attachment=False)
    except Exception as e:
        print(f"--- DEBUG ERROR: Erro ao servir ficheiro: {e} ---")
        abort(500)







# --- Rotas de Relatórios ---
@app.route('/api/relatorios/kpis')
@login_required
@gerente_required
def api_relatorios_kpis():
    """API para buscar KPIs dos relatórios."""
    conn = get_db()
    cur = conn.cursor(row_factory=dict_row)
    
    # Estatísticas gerais
    cur.execute('SELECT COUNT(*) as count FROM clientes')
    total_clientes = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM cobrancas WHERE status = 'Pendente'")
    cobrancas_pendentes = cur.fetchone()['count']
    
    cur.execute("SELECT COUNT(*) as count FROM cobrancas WHERE status = 'Pendente' AND data_vencimento < CURRENT_DATE")
    cobrancas_vencidas = cur.fetchone()['count']
    
    # Contagem correta: Número de parcelas que foram individualmente pagas
    cur.execute("SELECT COUNT(*) as count FROM parcelas WHERE status = 'Pago'")
    parcelas_pagas = cur.fetchone()['count']
    
    # Nova lógica de cálculo baseada no Saldo Devedor Total
    from datetime import date
    hoje = date.today()
    primeiro_dia_mes = hoje.replace(day=1)
    
    # Buscar todas as cobranças não pagas para calcular o saldo devedor total
    cur.execute("SELECT * FROM cobrancas WHERE status = 'Pendente'")
    cobrancas_abertas = cur.fetchall()
    
    saldo_devedor_total = 0
    for cobranca in cobrancas_abertas:
        # Buscar parcelas pendentes desta cobrança
        cur.execute('''
            SELECT * FROM parcelas 
            WHERE cobranca_id = %s AND status = 'Pendente'
        ''', (cobranca['id'],))
        parcelas_pendentes = cur.fetchall()
        
        for parcela in parcelas_pendentes:
            valor_parcela_atualizado = parcela['valor'] + (parcela['multa_manual'] or 0)
            saldo_devedor_total += valor_parcela_atualizado
    
    # Calcular total recebido no mês atual
    # Usa historico_pagamentos pois é o log unificado de todos os pagamentos (avulsos e parcelas)
    cur.execute("SELECT COALESCE(SUM(valor_pago), 0) as total FROM historico_pagamentos WHERE DATE(data_pagamento) >= %s", (primeiro_dia_mes,))
    total_recebido_mes = cur.fetchone()['total']
    
    cur.close()
    conn.close()
    
    return jsonify({
        'total_clientes': total_clientes,
        'cobrancas_pendentes': cobrancas_pendentes,
        'cobrancas_vencidas': cobrancas_vencidas,
        'parcelas_pagas': parcelas_pagas,
        'saldo_devedor_total': saldo_devedor_total,
        'total_recebido_mes': total_recebido_mes,
    })

@app.route('/relatorios')
@login_required
@gerente_required
def relatorios():
    """Página principal de relatórios."""
    return render_template('relatorios.html')





@app.route('/relatorios/clientes', methods=['GET', 'POST'])
@login_required
@gerente_required
def gerar_relatorio_clientes():
    """Gera relatório de clientes em Excel."""
    try:
        conn = get_db()
        cur = conn.cursor(row_factory=dict_row)
        
        # Buscar todos os clientes
        cur.execute('''
            SELECT 
                c.id,
                c.nome,
                c.cpf_cnpj,
                c.telefone,
                c.email,
                c.cidade,
                c.criado_em
            FROM clientes c
            ORDER BY c.nome
        ''')
        clientes = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Clientes"

        # --- Estilo do Cabeçalho ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")

        # --- Escreve o Cabeçalho ---
        headers = ['ID', 'Nome', 'CPF/CNPJ', 'Telefone', 'Email', 'Cidade', 'Data de Cadastro']
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # --- Escreve os Dados dos Clientes ---
        for cliente in clientes:
            ws.append([
                cliente['id'],
                cliente['nome'],
                cliente['cpf_cnpj'],
                cliente['telefone'],
                cliente['email'] or 'N/A',
                cliente['cidade'],
                cliente['criado_em'].strftime('%d/%m/%Y') if cliente['criado_em'] else 'N/A'
            ])

        # --- Autoajuste da Largura das Colunas ---
        for col_num, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].auto_size = True

        # --- Salva em Memória e Envia como Resposta ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=relatorio_clientes.xlsx"}
        )
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('relatorios'))

@app.route('/relatorios/cobrancas', methods=['GET', 'POST'])
@login_required
@gerente_required
def gerar_relatorio_cobrancas():
    """Gera relatório de cobranças em Excel."""
    try:
        conn = get_db()
        cur = conn.cursor(row_factory=dict_row)
        
        # Buscar todas as cobranças com dados do cliente
        cur.execute('''
            SELECT 
                c.id,
                cl.nome,
                c.valor_original,
                c.valor_total,
                c.valor_pago,
                c.data_vencimento,
                c.status,
                c.data_pagamento
            FROM cobrancas c
            JOIN clientes cl ON c.cliente_id = cl.id
            ORDER BY c.data_vencimento DESC
        ''')
        cobrancas = cur.fetchall()
        
        cur.close()
        conn.close()
        
        # Criar workbook Excel
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Cobranças"

        # --- Estilo do Cabeçalho ---
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="C0504D", end_color="C0504D", fill_type="solid")

        # --- CABEÇALHO ATUALIZADO ---
        headers = [
            'ID Cobrança', 'Nome Cliente', 'Valor Original', 'Valor Devido (c/ juros)', 
            'Total Pago', 'Nº de Pagamentos', 'Data Vencimento', 'Status Pagamento'
        ]
        ws.append(headers)

        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.font = header_font
            cell.fill = header_fill

        # --- Escreve os Dados das Cobranças ---
        for cobranca in cobrancas:
            # LÓGICA ATUALIZADA PARA INCLUIR DADOS DE PAGAMENTO
            total_pago = cobranca['valor_pago'] or 0.0
            
            # Buscar número de pagamentos individuais
            conn_temp = get_db()
            cur_temp = conn_temp.cursor(row_factory=dict_row)
            cur_temp.execute('SELECT COUNT(*) as count FROM pagamentos WHERE cobranca_id = %s', (cobranca['id'],))
            numero_pagamentos = cur_temp.fetchone()['count']
            cur_temp.close()
            conn_temp.close()

            ws.append([
                cobranca['id'], 
                cobranca['nome'], 
                cobranca['valor_original'],
                cobranca['valor_total'] or cobranca['valor_original'],
                total_pago,
                numero_pagamentos,
                cobranca['data_vencimento'].strftime('%d/%m/%Y') if cobranca['data_vencimento'] else 'N/A',
                'Pago' if cobranca['status'] == 'Pago' else 'Pendente'
            ])

        # --- Autoajuste da Largura das Colunas ---
        for col_num, _ in enumerate(headers, 1):
            column_letter = get_column_letter(col_num)
            ws.column_dimensions[column_letter].auto_size = True

        # --- Salva em Memória e Envia como Resposta ---
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)

        return Response(
            output,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment;filename=relatorio_cobrancas.xlsx"}
        )
        
    except Exception as e:
        flash(f'Erro ao gerar relatório: {str(e)}', 'danger')
        return redirect(url_for('relatorios'))

# --- Funções para Templates ---
@app.context_processor
def utility_processor():
    def get_user_nivel():
        """Retorna o nível do usuário logado."""
        return session.get('usuario_nivel', 'Operador')
    
    def can_access_reports():
        """Verifica se o usuário pode acessar relatórios."""
        return session.get('usuario_nivel') in ['Gerente', 'ADM']
    
    def can_access_admin():
        """Verifica se o usuário pode acessar funcionalidades administrativas."""
        return session.get('usuario_nivel') == 'ADM'
    
    return dict(
        get_user_nivel=get_user_nivel,
        can_access_reports=can_access_reports,
        can_access_admin=can_access_admin
    )

# --- Execução da Aplicação ---
if __name__ == '__main__':
    # Verificar se DATABASE_URL está configurada
    if not DATABASE_URL:
        print("❌ ERRO: Variável de ambiente DATABASE_URL não está configurada!")
        print("Configure a DATABASE_URL antes de executar o app.")
        print("Exemplo: export DATABASE_URL='postgresql://usuario:senha@localhost:5432/crm_db'")
        print("\nOu execute primeiro: python init_db.py")
        sys.exit(1)

    try:
        init_db()  # Inicializa o banco na primeira execução
        print("✅ Banco de dados inicializado com sucesso!")
    except Exception as e:
        print(f"❌ Erro ao inicializar banco de dados: {e}")
        print("Execute primeiro: python init_db.py")
        sys.exit(1)

    print("🚀 Iniciando aplicação Flask...")
    port = int(os.environ.get("PORT", 5000))  # Pega a porta do ambiente ou usa 5000 como padrão
    app.run(host='0.0.0.0', port=port, debug=True)
